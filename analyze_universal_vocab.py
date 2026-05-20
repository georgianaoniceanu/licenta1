import openpyxl
import pandas as pd
import numpy as np

# Load Excel file
wb = openpyxl.load_workbook('articles/lemmas_60k_subgenres.xlsx')
ws = wb['lemmas_60k_subgenres']
subcats = wb['sub-categories']

# Get subcategory mapping
subcat_dict = {}
for row in subcats.iter_rows(min_row=2, values_only=True):
    if row[0]:
        subcat_dict[f'x{row[0]}'] = row[1]
        subcat_dict[f'p{row[0]}'] = row[1]

print(f"Total subcategories: {len(subcat_dict)//2}")
print(f"Genres: {set(subcat_dict.values())}\n")

# Read all data into DataFrame
data = []
for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
    if i >= 100:  # First 100 words
        break
    rank, lemma, pos = row[0], row[1], row[2]
    frequencies = row[3:]
    data.append({
        'rank': rank,
        'lemma': lemma,
        'pos': pos,
        'frequencies': frequencies
    })

print(f"Loaded {len(data)} words\n")

# Calculate cross-domain utility
print("="*80)
print("TOP 30 UNIVERSAL VOCABULARY (highest frequency across ALL genres)")
print("="*80)

for item in data[:30]:
    frequencies = item['frequencies']
    # Get frequency percentages (every other value, starting from column index 97 for percentages)
    freq_counts = [f for f in frequencies[97:] if f is not None and isinstance(f, (int, float))]
    
    if freq_counts:
        mean_freq = np.mean(freq_counts)
        std_freq = np.std(freq_counts)
        min_freq = np.min(freq_counts)
        max_freq = np.max(freq_counts)
        cv = std_freq / mean_freq if mean_freq > 0 else 0  # Coefficient of variation
        
        print(f"\n#{item['rank']:4d} | {item['lemma']:12s} | PoS: {item['pos']}")
        print(f"       Mean: {mean_freq:8.2f} | Std: {std_freq:8.2f} | CV: {cv:.4f}")
        print(f"       Range: {min_freq:.2f} - {max_freq:.2f}")

print("\n" + "="*80)
print("EXPLANATION:")
print("- CV (Coefficient of Variation) < 0.5 = UNIVERSAL (consistent across genres)")
print("- CV > 0.8 = GENRE-SPECIFIC (varies greatly)")
print("="*80)
