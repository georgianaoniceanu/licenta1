import pdfplumber
import openpyxl
import os

# Extract PDFs
pdf_files = ['articles/amt018.pdf', 'articles/davies_52.pdf']

for pdf in pdf_files:
    print(f"\n{'='*60}")
    print(f"READING: {pdf}")
    print(f"{'='*60}")
    try:
        with pdfplumber.open(pdf) as doc:
            print(f"Total pages: {len(doc.pages)}")
            # Extract first 3 pages
            for i, page in enumerate(doc.pages[:3]):
                text = page.extract_text()
                print(f"\n--- PAGE {i+1} ---")
                print(text[:500] if text else "[No text extracted]")
    except Exception as e:
        print(f"Error: {e}")

# Read Excel file
print(f"\n{'='*60}")
print("READING: articles/lemmas_60k_subgenres.xlsx")
print(f"{'='*60}")
try:
    wb = openpyxl.load_workbook('articles/lemmas_60k_subgenres.xlsx')
    print(f"Sheet names: {wb.sheetnames}")
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\n--- SHEET: {sheet_name} ---")
        print(f"Dimensions: {ws.dimensions}")
        print(f"Max row: {ws.max_row}, Max col: {ws.max_column}")
        
        # Print first 10 rows
        print("\nFirst 10 rows:")
        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True)):
            print(f"  Row {i+1}: {row}")
            
except Exception as e:
    print(f"Error: {e}")
